"""维修管理"""
import io
import json
import logging
from datetime import datetime, date, timedelta
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc, func, distinct
from pydantic import BaseModel
from database import get_db
from models.repair import RepairOrder, RepairDetail, WarrantyAlert
from models.vehicle import Vehicle
from models.user import User
from services.auth import get_current_user, require_role
from services.operation_log import log_operation
from services.permissions import scope_query, get_descendant_group_ids
from models.user import Group

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/repairs", tags=["维修"])


# ---------- 修理厂列表 ----------
@router.get("/shops", include_in_schema=False)
async def list_shops(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """返回所有修理厂名称列表（用于下拉选择）"""
    # 从 repair_shop 角色的用户获取
    users_r = await db.execute(
        select(User).where(User.role == "repair_shop", User.is_active == True)
    )
    shops_from_users = [u.name for u in users_r.scalars().all() if u.name]
    # 从已有维修单提取不同的 shop_name
    orders_r = await db.execute(select(distinct(RepairOrder.shop_name)))
    shops_from_orders = [r[0] for r in orders_r if r[0]]
    # 合并去重
    all_shops = list(dict.fromkeys(shops_from_users + shops_from_orders))
    return {"shops": sorted(all_shops)}


class DispatchReq(BaseModel):
    """车管员派单"""
    vehicle_id: int
    description: str = ""
    shop_name: str = ""
    assigned_to: int | None = None
    dispatch_photos: str = "[]"
    flow_id: int | None = None


class AcceptReq(BaseModel):
    """汽修厂接单"""
    accept_photos: str = "[]"
    notes: str = ""


class DetailReq(BaseModel):
    """汽修厂上传维修明细"""
    items: list[dict]  # [{item_name, item_desc, parts_used, cost, warranty_days}]
    photos: str = "[]"


class VerifyReq(BaseModel):
    """车管员验收"""
    verify_photos: str = "[]"
    notes: str = ""


@router.post("/dispatch")
async def dispatch_repair(req: DispatchReq, db: AsyncSession = Depends(get_db),
                          user: User = Depends(require_role("fleet_manager", "admin"))):
    """车管员派单"""
    vehicle = await db.get(Vehicle, req.vehicle_id)
    if not vehicle:
        raise HTTPException(404, "车辆不存在")
    order = RepairOrder(
        vehicle_id=req.vehicle_id,
        created_by=user.id,
        assigned_to=req.assigned_to,
        shop_name=req.shop_name,
        description=req.description,
        dispatch_photos=req.dispatch_photos,
        flow_id=req.flow_id,
        current_step=1 if req.flow_id else 0,
        status="dispatched",
    )
    db.add(order)
    await db.commit()
    await db.refresh(order)
    plate = vehicle.plate_number if vehicle else ""
    await log_operation(db, user, f"派发了维修单 #{order.id}（{plate}→{order.shop_name}）", "repair", order.id)
    return {"id": order.id, "status": "dispatched"}


@router.get("")
async def list_repairs(
    status: str | None = None,
    page: int = 1,
    size: int = 20,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """维修单列表"""
    query = select(RepairOrder).order_by(desc(RepairOrder.id))
    count_query = select(func.count(RepairOrder.id)).select_from(RepairOrder)
    conds = []
    if status:
        conds.append(RepairOrder.status == status)
    query = await scope_query(query, user, RepairOrder)
    count_query = await scope_query(count_query, user, RepairOrder)

    total_r = await db.execute(count_query)
    total = total_r.scalar() or 0
    query = query.offset((page - 1) * size).limit(size)
    result = await db.execute(query)
    orders = result.scalars().all()

    ret = []
    for o in orders:
        vehicle = await db.get(Vehicle, o.vehicle_id)
        ret.append({
            "id": o.id,
            "vehicle_id": o.vehicle_id,
            "plate_number": vehicle.plate_number if vehicle else "",
            "shop_name": o.shop_name,
            "description": o.description,
            "status": o.status,
            "flow_id": o.flow_id,
            "current_step": o.current_step,
            "created_at": o.created_at.isoformat() if o.created_at else None,
            "detail_count": len(o.details),
        })
    return {"items": ret, "total": total, "page": page, "size": size}


@router.get("/{oid}")
async def get_repair(oid: int, db: AsyncSession = Depends(get_db),
                     user: User = Depends(get_current_user)):
    """维修单详情"""
    order = await db.get(RepairOrder, oid)
    if not order:
        raise HTTPException(404, "维修单不存在")
    vehicle = await db.get(Vehicle, order.vehicle_id)
    return {
        "id": order.id,
        "vehicle_id": order.vehicle_id,
        "plate_number": vehicle.plate_number if vehicle else "",
        "shop_name": order.shop_name,
        "description": order.description,
        "dispatch_photos": json.loads(order.dispatch_photos),
        "accept_photos": json.loads(order.accept_photos),
        "verify_photos": json.loads(order.verify_photos),
        "status": order.status,
        "flow_id": order.flow_id,
        "current_step": order.current_step,
        "verify_notes": order.verify_notes,
        "created_at": order.created_at.isoformat() if order.created_at else None,
        "details": [{
            "id": d.id, "item_name": d.item_name, "item_desc": d.item_desc,
            "parts_used": d.parts_used, "cost": d.cost,
            "warranty_days": d.warranty_days, "warranty_end": d.warranty_end,
        } for d in order.details],
    }


@router.put("/{oid}/accept")
async def accept_repair(oid: int, req: AcceptReq, db: AsyncSession = Depends(get_db),
                        user: User = Depends(require_role("repair_shop"))):
    """汽修厂接单"""
    order = await db.get(RepairOrder, oid)
    if not order:
        raise HTTPException(404, "维修单不存在")
    if order.status != "dispatched":
        raise HTTPException(400, "当前状态不可接单")
    order.status = "accepted"
    order.assigned_to = user.id
    order.accept_photos = req.accept_photos
    order.accepted_at = datetime.now()
    await db.commit()
    await log_operation(db, user, f"接单了维修单 #{oid}", "repair", oid)
    return {"msg": "已接单", "status": "accepted"}


@router.put("/{oid}/details")
async def upload_details(oid: int, req: DetailReq, db: AsyncSession = Depends(get_db),
                         user: User = Depends(require_role("repair_shop"))):
    """汽修厂上传维修明细"""
    order = await db.get(RepairOrder, oid)
    if not order:
        raise HTTPException(404, "维修单不存在")
    if order.status not in ("accepted", "in_progress"):
        raise HTTPException(400, "当前状态不可上传明细")

    order.status = "in_progress"

    # 删除旧明细
    for old in order.details:
        await db.delete(old)

    # 写入新明细
    now = date.today()
    for item in req.items:
        warranty_end = ""
        if item.get("warranty_days", 0) > 0:
            end = now + timedelta(days=item["warranty_days"])
            warranty_end = end.isoformat()

        detail = RepairDetail(
            order_id=oid,
            item_name=item["item_name"],
            item_desc=item.get("item_desc", ""),
            parts_used=item.get("parts_used", ""),
            cost=item.get("cost", 0),
            warranty_days=item.get("warranty_days", 0),
            warranty_end=warranty_end,
            photos=req.photos,
        )
        db.add(detail)

        # 检查质保期内再次维修
        await _check_warranty(db, oid, order.vehicle_id, item["item_name"], warranty_end)

    await db.commit()
    await log_operation(db, user, f"提交了维修单 #{oid} 的维修明细（{len(req.items)}项）", "repair", oid)
    return {"msg": "明细已上传"}


async def _check_warranty(db, new_order_id, vehicle_id, item_name, warranty_end):
    """检查该车辆同一项目是否有在质保期内的旧维修记录"""
    result = await db.execute(
        select(RepairDetail).where(
            RepairDetail.item_name == item_name,
            RepairDetail.warranty_end.isnot(None),
            RepairDetail.warranty_end != "",
        ).order_by(desc(RepairDetail.id))
    )
    old_details = result.scalars().all()
    for od in old_details:
        # 检查是否同一车辆
        old_order = await db.get(RepairOrder, od.order_id)
        if old_order and old_order.vehicle_id == vehicle_id:
            try:
                end_date = datetime.strptime(od.warranty_end, "%Y-%m-%d").date()
                if end_date >= date.today():
                    # 在质保期内，触发提醒
                    alert = WarrantyAlert(
                        detail_id=od.id,
                        vehicle_id=vehicle_id,
                        original_item=od.item_name,
                        warranty_end=od.warranty_end,
                        new_order_id=new_order_id,
                    )
                    db.add(alert)
                    logger.info(f"质保期内再次维修提醒: {item_name}, 车辆 {vehicle_id}")
            except ValueError:
                continue


@router.put("/{oid}/complete")
async def complete_repair(oid: int, db: AsyncSession = Depends(get_db),
                          user: User = Depends(require_role("repair_shop"))):
    """汽修厂标记维修完成"""
    order = await db.get(RepairOrder, oid)
    if not order:
        raise HTTPException(404)
    if order.status != "in_progress":
        raise HTTPException(400, "当前状态不可完成")
    order.status = "completed"
    order.completed_at = datetime.now()
    await db.commit()
    await log_operation(db, user, f"完成了维修单 #{oid}", "repair", oid)
    return {"msg": "已完成", "status": "completed"}


@router.put("/{oid}/verify")
async def verify_repair(oid: int, req: VerifyReq, db: AsyncSession = Depends(get_db),
                        user: User = Depends(require_role("fleet_manager", "admin"))):
    """车管员验收"""
    order = await db.get(RepairOrder, oid)
    if not order:
        raise HTTPException(404)
    if order.status != "completed":
        raise HTTPException(400, "维修未完成，不可验收")
    order.status = "verified"
    order.verify_photos = req.verify_photos
    order.verify_notes = req.notes
    order.verified_at = datetime.now()
    await db.commit()
    await log_operation(db, user, f"验收了维修单 #{oid}", "repair", oid)
    return {"msg": "已验收", "status": "verified"}


@router.get("/warranty-alerts")
async def list_warranty_alerts(db: AsyncSession = Depends(get_db),
                               user: User = Depends(get_current_user)):
    """质保期提醒列表"""
    query = select(WarrantyAlert).order_by(desc(WarrantyAlert.id))
    if user.role != "admin":
        # 质保预警按车辆分组过滤
        _scope = "all"
        if user.role_id and user.role_obj:
            _scope = user.role_obj.data_scope or "all"
        elif user.role in ("fleet_manager",):
            _scope = "group"
        if _scope == "all":
            pass  # 看全部
        elif _scope == "group" and user.group_id:
            gids = await get_descendant_group_ids(db, user.group_id)
            vehs = await db.execute(select(Vehicle.id).where(Vehicle.group_id.in_(gids)))
            vids = [r[0] for r in vehs]
            if vids:
                query = query.where(WarrantyAlert.vehicle_id.in_(vids))
            else:
                query = query.where(1 == 0)
    result = await db.execute(query)
    alerts = result.scalars().all()
    return [{
        "id": a.id,
        "vehicle_id": a.vehicle_id,
        "original_item": a.original_item,
        "warranty_end": a.warranty_end,
        "new_order_id": a.new_order_id,
        "is_read": a.is_read,
        "created_at": a.created_at.isoformat() if a.created_at else None,
    } for a in alerts]


@router.put("/warranty-alerts/{aid}/read")
async def mark_alert_read(aid: int, db: AsyncSession = Depends(get_db),
                          user: User = Depends(get_current_user)):
    alert = await db.get(WarrantyAlert, aid)
    if not alert:
        raise HTTPException(404)
    alert.is_read = True
    await db.commit()
    return {"msg": "已读"}


@router.get("/export")
async def export_repairs(
    status: str | None = None,
    shop_name: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """导出维修单 Excel（支持按修理厂筛选）"""
    query = select(RepairOrder).order_by(desc(RepairOrder.id))
    if status:
        query = query.where(RepairOrder.status == status)
    if shop_name:
        query = query.where(RepairOrder.shop_name.ilike(f"%{shop_name}%"))
    query = await scope_query(query, user, RepairOrder)

    result = await db.execute(query)
    orders = result.scalars().all()

    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "维修记录"

    # 表头
    headers = ["单号", "车牌号", "修理厂", "描述", "状态", "总费用", "派单时间", "完成时间"]
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    status_map = {"dispatched": "已派单", "accepted": "已接单", "submitted": "已提交明细",
                  "in_progress": "维修中", "completed": "已完成", "verified": "已验收", "rejected": "已驳回"}

    for i, o in enumerate(orders, 2):
        vehicle = await db.get(Vehicle, o.vehicle_id)
        plate = vehicle.plate_number if vehicle else ""
        row_data = [
            o.id, plate, o.shop_name or "", o.description or "",
            status_map.get(o.status, o.status), o.total_cost or "",
            o.created_at.isoformat() if o.created_at else "",
            o.completed_at.isoformat() if o.completed_at else "",
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # 自动列宽
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(content=buf.getvalue(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=repairs.xlsx"})


@router.post("/{oid}/advance")
async def advance_repair_step(
    oid: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """自定义流程：推进到下一步（仅当前步骤的操作角色可执行）"""
    order = await db.get(RepairOrder, oid)
    if not order:
        raise HTTPException(404, "维修单不存在")
    if not order.flow_id:
        raise HTTPException(400, "此维修单使用默认流程，请使用对应操作按钮")

    # 获取当前步骤
    steps_r = await db.execute(
        select(RepairFlowStep).where(RepairFlowStep.flow_id == order.flow_id)
        .order_by(RepairFlowStep.step_order)
    )
    steps = steps_r.scalars().all()
    if not steps:
        raise HTTPException(400, "流程未配置步骤")

    current_idx = order.current_step - 1  # steps 列表索引
    if current_idx < 0 or current_idx >= len(steps):
        raise HTTPException(400, "维修单已结束或状态异常")

    step = steps[current_idx]
    # 检查角色权限
    if step.action_role and user.role != step.action_role:
        raise HTTPException(403, f"当前步骤【{step.step_name}】需要【{step.action_role}】角色操作")

    # 推进到下一步
    next_step = current_idx + 1
    if next_step >= len(steps):
        # 全部步骤完成
        order.current_step = len(steps) + 1
        order.status = "completed"
        order.completed_at = datetime.now()
        await db.commit()
        await log_operation(db, user, f"完成了维修单 #{oid}（全部流程步骤已完成）", "repair", oid)
        return {"status": "completed", "msg": "全部流程步骤已完成"}
    else:
        order.current_step = next_step + 1
        order.status = f"step_{next_step + 1}"
        await db.commit()
        next_s = steps[next_step]
        await log_operation(db, user, f"推进维修单 #{oid} 到【{next_s.step_name}】", "repair", oid)
        return {
            "status": f"step_{next_step + 1}",
            "current_step": order.current_step,
            "step_name": next_s.step_name,
            "action_role": next_s.action_role,
            "action_label": next_s.action_label,
            "msg": f"已推进到【{next_s.step_name}】",
        }
