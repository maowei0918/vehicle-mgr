"""车辆管理"""
import csv
import io
import logging
from datetime import date
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import Response
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from pydantic import BaseModel
from database import get_db
from models.vehicle import Vehicle
from models.user import User
from services.auth import get_current_user, require_role
from services.operation_log import log_operation
from services.permissions import scope_query

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/vehicles", tags=["车辆"])


class VehicleReq(BaseModel):
    plate_number: str
    brand: str = ""
    model: str = ""
    color: str = ""
    vin: str = ""
    registration_date: str = ""
    group_id: int | None = None
    driver_id: int | None = None
    notes: str = ""


class VehicleImportItem(BaseModel):
    plate_number: str
    model: str = ""
    registration_date: str = ""
    driver_username: str = ""


@router.get("/template.xlsx")
async def download_vehicle_template():
    """下载车辆导入模板（xlsx格式）"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "服务器未安装 openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "车辆导入模板"
    headers = ["车牌号", "车型", "注册日期", "责任人(用户名)"]
    ws.append(headers)
    ws.append(["京A12345", "大众迈腾 2024款", "2024-06-01", "zhangsan"])
    ws.append(["京B67890", "丰田凯美瑞 2023款", "2023-09-15", "lisi"])
    # 设置列宽
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=vehicle-import-template.xlsx"},
    )


@router.get("")
async def list_vehicles(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    query = select(Vehicle).order_by(Vehicle.id)
    query = await scope_query(query, user, Vehicle)
    result = await db.execute(query)
    vehicles = result.scalars().all()
    return [{
        "id": v.id, "plate_number": v.plate_number, "brand": v.brand,
        "model": v.model, "color": v.color, "vin": v.vin,
        "registration_date": v.registration_date.isoformat() if v.registration_date else None,
        "group_id": v.group_id, "driver_id": v.driver_id,
        "driver_name": v.driver.name if v.driver else None,
        "is_active": v.is_active, "notes": v.notes,
    } for v in vehicles]


@router.post("")
async def create_vehicle(req: VehicleReq, db: AsyncSession = Depends(get_db),
                         user: User = Depends(require_role("admin", "fleet_manager"))):
    exist = await db.execute(select(Vehicle).where(Vehicle.plate_number == req.plate_number))
    if exist.first():
        raise HTTPException(400, "车牌号已存在")
    data = req.model_dump()
    if data.get("registration_date"):
        try:
            data["registration_date"] = date.fromisoformat(data["registration_date"])
        except ValueError:
            data["registration_date"] = None
    v = Vehicle(**data)
    if user.role == "fleet_manager":
        v.group_id = user.group_id
    db.add(v)
    await db.commit()
    await db.refresh(v)
    await log_operation(db, user, f"添加了车辆「{v.plate_number}」", "vehicle", v.id)
    return {"id": v.id, "plate_number": v.plate_number}


@router.put("/{vid}")
async def update_vehicle(vid: int, req: VehicleReq, db: AsyncSession = Depends(get_db),
                         user: User = Depends(require_role("admin", "fleet_manager"))):
    v = await db.get(Vehicle, vid)
    if not v:
        raise HTTPException(404, "车辆不存在")
    data = req.model_dump()
    if data.get("registration_date"):
        try:
            data["registration_date"] = date.fromisoformat(data["registration_date"])
        except ValueError:
            data["registration_date"] = None
    for key, val in data.items():
        setattr(v, key, val)
    await db.commit()
    await log_operation(db, user, f"编辑了车辆「{v.plate_number}」", "vehicle", vid)
    return {"msg": "已更新"}


@router.post("/import")
async def import_vehicles(items: list[VehicleImportItem],
                          db: AsyncSession = Depends(get_db),
                          user: User = Depends(require_role("admin", "fleet_manager"))):
    """批量导入车辆（格式：车牌号,车型,注册日期,责任人用户名）"""
    imported = 0
    skipped = 0
    for item in items:
        if not item.plate_number.strip():
            skipped += 1
            continue
        exist = await db.execute(select(Vehicle).where(Vehicle.plate_number == item.plate_number.strip()))
        if exist.first():
            skipped += 1
            continue
        driver_id = None
        if item.driver_username.strip():
            driver = await db.execute(select(User).where(User.username == item.driver_username.strip()))
            driver_user = driver.scalar_one_or_none()
            if driver_user:
                driver_id = driver_user.id
        reg_date = None
        if item.registration_date.strip():
            try:
                reg_date = date.fromisoformat(item.registration_date.strip())
            except ValueError:
                pass
        v = Vehicle(
            plate_number=item.plate_number.strip(),
            model=item.model.strip() or "",
            registration_date=reg_date,
            driver_id=driver_id,
        )
        if user.role == "fleet_manager":
            v.group_id = user.group_id
        db.add(v)
        imported += 1
    await db.commit()
    await log_operation(db, user, f"批量导入了 {imported} 辆车辆（跳过 {skipped} 个）", "vehicle", 0)
    return {"imported": imported, "skipped": skipped}


@router.get("/export")
async def export_vehicles(
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """导出车辆 CSV"""
    query = select(Vehicle).order_by(Vehicle.id)
    query = await scope_query(query, user, Vehicle)
    result = await db.execute(query)
    vehicles = result.scalars().all()

    import csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow(["车牌号", "品牌", "车型", "颜色", "车架号", "注册日期", "责任人", "分组ID", "备注"])
    for v in vehicles:
        driver_name = v.driver.name if v.driver else ""
        reg_date = v.registration_date.isoformat() if v.registration_date else ""
        w.writerow([v.plate_number, v.brand, v.model, v.color, v.vin, reg_date, driver_name, v.group_id or "", v.notes or ""])
    return Response(content=buf.getvalue(), media_type="text/csv; charset=utf-8",
                    headers={"Content-Disposition": "attachment; filename=vehicles.csv"})


@router.post("/import/xlsx")
async def import_vehicles_xlsx(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_role("admin", "fleet_manager")),
):
    """从 xlsx 导入车辆"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "服务器未安装 openpyxl")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    if ws is None:
        raise HTTPException(400, "Excel 文件为空")

    imported = 0
    skipped = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            skipped += 1
            continue
        # 列: 车牌号, 车型, 注册日期, 责任人(用户名)
        plate_number = str(row[0] or "").strip()
        if not plate_number:
            skipped += 1
            continue
        exist = await db.execute(select(Vehicle).where(Vehicle.plate_number == plate_number))
        if exist.first():
            skipped += 1
            continue
        vehicle_model = str(row[1] or "").strip() if len(row) > 1 else ""
        reg_date_str = str(row[2] or "").strip() if len(row) > 2 else ""
        driver_username = str(row[3] or "").strip() if len(row) > 3 else ""

        driver_id = None
        if driver_username:
            driver = await db.execute(select(User).where(User.username == driver_username))
            driver_user = driver.scalar_one_or_none()
            if driver_user:
                driver_id = driver_user.id
        reg_date = None
        if reg_date_str:
            try:
                reg_date = date.fromisoformat(reg_date_str)
            except ValueError:
                pass
        v = Vehicle(
            plate_number=plate_number,
            model=vehicle_model,
            registration_date=reg_date,
            driver_id=driver_id,
        )
        if user.role == "fleet_manager":
            v.group_id = user.group_id
        db.add(v)
        imported += 1
    await db.commit()
    await log_operation(db, user, f"批量导入了 {imported} 辆车辆（跳过 {skipped} 个）", "vehicle", 0)
    return {"imported": imported, "skipped": skipped}
