"""分组 & 用户管理"""
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from sqlalchemy.orm import selectinload
from pydantic import BaseModel
from database import get_db
from models.user import Group, User
from models.role import Role
from services.auth import hash_password, get_current_user, require_role
from services.operation_log import log_operation
from services.permissions import scope_query

router = APIRouter(prefix="/api", tags=["分组&用户"])


# ========== 分组 ==========

class GroupReq(BaseModel):
    name: str = ""
    desc: str = ""
    parent_id: int | None = None


@router.get("/groups")
async def list_groups(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    result = await db.execute(
        select(Group).options(selectinload(Group.children), selectinload(Group.users)).order_by(Group.id)
    )
    groups = result.scalars().all()
    return [{
        "id": g.id, "name": g.name, "desc": g.desc,
        "parent_id": g.parent_id,
        "member_count": len(g.users),
        "children": [c.id for c in g.children] if g.children else [],
    } for g in groups]


@router.post("/groups")
async def create_group(req: GroupReq, db: AsyncSession = Depends(get_db),
                       user: User = Depends(require_role("admin"))):
    if not req.name.strip():
        raise HTTPException(400, "名称不能为空")
    if req.parent_id:
        parent = await db.get(Group, req.parent_id)
        if not parent:
            raise HTTPException(400, "上级分组不存在")
    g = Group(name=req.name.strip(), desc=req.desc.strip(), parent_id=req.parent_id)
    db.add(g)
    await db.commit()
    await db.refresh(g)
    await log_operation(db, user, f"创建了分组「{g.name}」", "group", g.id)
    return {"id": g.id, "name": g.name, "desc": g.desc, "parent_id": g.parent_id}


@router.put("/groups/{gid}")
async def update_group(gid: int, req: GroupReq, db: AsyncSession = Depends(get_db),
                       user: User = Depends(require_role("admin"))):
    g = await db.get(Group, gid)
    if not g:
        raise HTTPException(404, "分组不存在")
    if req.name.strip():
        g.name = req.name.strip()
    if req.desc is not None:
        g.desc = req.desc.strip()
    # 允许重新指定 parent_id，但不能设为自己的子分组（防环）
    if "parent_id" in req.model_dump(exclude_unset=True):
        if req.parent_id == gid:
            raise HTTPException(400, "不能将自己设为上级分组")
        g.parent_id = req.parent_id
    await db.commit()
    await log_operation(db, user, f"编辑了分组「{g.name}」", "group", gid)
    return {"id": g.id, "name": g.name, "desc": g.desc, "parent_id": g.parent_id}


@router.delete("/groups/{gid}")
async def delete_group(gid: int, db: AsyncSession = Depends(get_db),
                       user: User = Depends(require_role("admin"))):
    g = await db.get(Group, gid)
    if not g:
        raise HTTPException(404, "分组不存在")
    if g.users:
        raise HTTPException(400, "分组下还有成员，无法删除")
    if g.children:
        raise HTTPException(400, "分组下还有子分组，请先删除子分组")
    await log_operation(db, user, f"删除了分组「{g.name}」", "group", gid)
    await db.delete(g)
    await db.commit()
    return {"msg": "已删除"}


# ========== 用户 ==========

class UserReq(BaseModel):
    username: str = ""
    password: str = ""
    name: str = ""
    phone: str = ""
    role: str = "driver"
    role_id: int | None = None
    group_id: int | None = None


@router.get("/users")
async def list_users(db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    query = select(User).options(selectinload(User.group), selectinload(User.role_obj)).order_by(User.id)
    query = await scope_query(query, user, User)
    result = await db.execute(query)
    users = result.scalars().all()
    return [{
        "id": u.id, "username": u.username, "name": u.name, "phone": u.phone,
        "role": u.role, "role_id": u.role_id, "role_name": u.role_obj.display_name if u.role_obj else u.role,
        "group_id": u.group_id, "is_active": u.is_active,
    } for u in users]


@router.post("/users")
async def create_user(req: UserReq, db: AsyncSession = Depends(get_db),
                      user: User = Depends(require_role("admin", "fleet_manager"))):
    if not req.username:
        raise HTTPException(400, "用户名不能为空")
    if not req.name:
        req.name = req.username
    if user.role == "fleet_manager":
        req.group_id = user.group_id  # 只能创建自己组的成员
    if req.role not in ("admin", "fleet_manager", "repair_shop", "driver"):
        req.role = "driver"
    # 如果传了 role_id 则根据角色对象设 role
    if req.role_id is not None:
        role_obj = await db.get(Role, req.role_id)
        if role_obj:
            req.role = role_obj.name
    exist = await db.execute(select(User).where(User.username == req.username))
    if exist.first():
        raise HTTPException(400, "用户名已存在")
    u = User(
        username=req.username, password_hash=hash_password(req.password or "123456"),
        name=req.name, phone=req.phone, role=req.role, group_id=req.group_id,
        role_id=req.role_id,
    )
    db.add(u)
    await db.commit()
    await db.refresh(u)
    await log_operation(db, user, f"创建了用户「{u.name}」({u.username})", "user", u.id)
    return {"id": u.id, "name": u.name, "role": u.role}


@router.put("/users/{uid}")
async def update_user(uid: int, req: UserReq, db: AsyncSession = Depends(get_db),
                      user: User = Depends(require_role("admin", "fleet_manager"))):
    u = await db.get(User, uid)
    if not u:
        raise HTTPException(404, "用户不存在")
    if user.role == "fleet_manager" and u.group_id != user.group_id:
        raise HTTPException(403, "只能管理自己分组的用户")
    if req.name:
        u.name = req.name
    if req.phone:
        u.phone = req.phone
    if req.password:
        u.password_hash = hash_password(req.password)
    if req.role and user.role == "admin":
        u.role = req.role
    if req.role_id is not None and user.role == "admin":
        u.role_id = req.role_id
    if req.group_id is not None and user.role == "admin":
        u.group_id = req.group_id
    await db.commit()
    await log_operation(db, user, f"编辑了用户「{u.name}」({u.username})", "user", uid)
    return {"msg": "已更新"}


@router.delete("/users/{uid}")
async def delete_user(uid: int, db: AsyncSession = Depends(get_db),
                      user: User = Depends(require_role("admin"))):
    """删除用户（仅管理员）"""
    u = await db.get(User, uid)
    if not u:
        raise HTTPException(404, "用户不存在")
    if u.role == "admin":
        # 检查是否最后一个管理员
        result = await db.execute(select(User).where(User.role == "admin"))
        admins = result.scalars().all()
        if len(admins) <= 1:
            raise HTTPException(400, "至少保留一个管理员账号")
    await log_operation(db, user, f"删除了用户「{u.name}」({u.username})", "user", uid)
    await db.delete(u)
    await db.commit()
    return {"msg": "已删除"}


class UserImportItem(BaseModel):
    username: str
    name: str = ""
    password: str = "123456"
    role: str = "driver"
    group_name: str = ""
    phone: str = ""


@router.post("/users/import")
async def import_users(items: list[UserImportItem],
                       db: AsyncSession = Depends(get_db),
                       user: User = Depends(require_role("admin", "fleet_manager"))):
    """批量导入用户（文本格式：用户名,姓名,密码,角色,分组名,手机号）"""
    # 读取所有分组名到 id 映射
    groups_r = await db.execute(select(Group))
    group_map = {g.name: g.id for g in groups_r.scalars().all()}

    imported = 0
    skipped = 0
    for item in items:
        if not item.username.strip():
            skipped += 1
            continue
        exist = await db.execute(select(User).where(User.username == item.username.strip()))
        if exist.first():
            skipped += 1
            continue
        role = item.role.strip() or "driver"
        if role not in ("admin", "fleet_manager", "repair_shop", "driver"):
            role = "driver"
        group_id = group_map.get(item.group_name.strip()) if item.group_name.strip() else None
        if user.role == "fleet_manager":
            group_id = user.group_id
        u = User(
            username=item.username.strip(),
            password_hash=hash_password(item.password.strip() or "123456"),
            name=item.name.strip() or item.username.strip(),
            phone=item.phone.strip() or "",
            role=role,
            group_id=group_id,
        )
        db.add(u)
        imported += 1
    await db.commit()
    await log_operation(db, user, f"批量导入了 {imported} 个用户（跳过 {skipped} 个）", "user", 0)
    return {"imported": imported, "skipped": skipped}


@router.get("/users/template.xlsx")
async def download_user_template():
    """下载用户导入模板（xlsx格式）"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "服务器未安装 openpyxl")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "用户导入模板"
    headers = ["用户名", "姓名", "密码", "角色", "分组", "手机号"]
    ws.append(headers)
    ws.append(["zhangsan", "张三", "123456", "driver", "第一车队", "13800138001"])
    ws.append(["lisi", "李四", "123456", "repair_shop", "", "13800138002"])
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import Response
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=user-import-template.xlsx"},
    )


@router.post("/users/import/xlsx")
async def import_users_xlsx(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_role("admin", "fleet_manager")),
):
    """从 xlsx 导入用户"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "服务器未安装 openpyxl")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    if ws is None:
        raise HTTPException(400, "Excel 文件为空")

    # 读取所有分组名到 id 映射
    groups_r = await db.execute(select(Group))
    group_map = {g.name: g.id for g in groups_r.scalars().all()}

    imported = 0
    skipped = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            skipped += 1
            continue
        # 列: 用户名, 姓名, 密码, 角色, 分组, 手机号
        username = str(row[0] or "").strip()
        name = str(row[1] or row[0] or "").strip()
        password = str(row[2] or "123456").strip()
        role = str(row[3] or "driver").strip()
        group_name = str(row[4] or "").strip() if len(row) > 4 else ""
        phone = str(row[5] or "").strip() if len(row) > 5 else ""

        if not username:
            skipped += 1
            continue
        if role not in ("admin", "fleet_manager", "repair_shop", "driver"):
            role = "driver"

        exist = await db.execute(select(User).where(User.username == username))
        if exist.first():
            skipped += 1
            continue

        group_id = group_map.get(group_name) if group_name else None
        if user.role == "fleet_manager":
            group_id = user.group_id

        u = User(
            username=username,
            password_hash=hash_password(password),
            name=name,
            phone=phone,
            role=role,
            group_id=group_id,
        )
        db.add(u)
        imported += 1
    await db.commit()
    await log_operation(db, user, f"批量导入了 {imported} 个用户（跳过 {skipped} 个）", "user", 0)
    return {"imported": imported, "skipped": skipped}

