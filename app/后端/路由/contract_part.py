"""合同配件管理"""
import io
import logging
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, desc, or_
from pydantic import BaseModel
from database import get_db
from models.contract_part import ContractPart
from models.user import User
from services.auth import get_current_user, require_role
from services.operation_log import log_operation

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/contract-parts", tags=["合同配件"])


class ContractPartReq(BaseModel):
    contract_id: int | None = None
    shop_id: int | None = None
    vehicle_model: str = ""
    part_name: str
    part_model: str = ""
    contract_price: float = 0.0
    labor_fee: float = 0.0
    total_amount: float = 0.0
    warranty_days: int = 0
    notes: str = ""


@router.get("")
async def list_contract_parts(
    q: str = "",
    contract_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    query = select(ContractPart).order_by(desc(ContractPart.id))
    if contract_id is not None:
        query = query.where(ContractPart.contract_id == contract_id)
    if q:
        query = query.where(
            or_(
                ContractPart.part_name.ilike(f"%{q}%"),
                ContractPart.vehicle_model.ilike(f"%{q}%"),
                ContractPart.part_model.ilike(f"%{q}%"),
            )
        )
    result = await db.execute(query)
    parts = result.scalars().all()
    return [
        {
            "id": p.id,
            "contract_id": p.contract_id,
            "shop_id": p.shop_id,
            "vehicle_model": p.vehicle_model,
            "part_name": p.part_name,
            "part_model": p.part_model,
            "contract_price": p.contract_price,
            "labor_fee": p.labor_fee,
            "total_amount": p.total_amount,
            "warranty_days": p.warranty_days,
            "notes": p.notes,
            "created_at": p.created_at.isoformat() if p.created_at else None,
        }
        for p in parts
    ]


@router.post("")
async def create_contract_part(
    req: ContractPartReq,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_role("admin")),
):
    if not req.part_name.strip():
        raise HTTPException(400, "配件名称不能为空")
    p = ContractPart(**req.model_dump())
    db.add(p)
    await db.commit()
    await db.refresh(p)
    await log_operation(db, user, f"添加了合同配件「{p.part_name}」（{p.vehicle_model}）", "contract_part", p.id)
    return {"id": p.id, "part_name": p.part_name}


@router.put("/{pid}")
async def update_contract_part(
    pid: int,
    req: ContractPartReq,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_role("admin")),
):
    p = await db.get(ContractPart, pid)
    if not p:
        raise HTTPException(404, "合同配件不存在")
    for key, val in req.model_dump().items():
        setattr(p, key, val)
    await db.commit()
    await log_operation(db, user, f"编辑了合同配件「{p.part_name}」", "contract_part", pid)
    return {"msg": "已更新"}


@router.delete("/{pid}")
async def delete_contract_part(
    pid: int,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_role("admin")),
):
    p = await db.get(ContractPart, pid)
    if not p:
        raise HTTPException(404, "合同配件不存在")
    await log_operation(db, user, f"删除了合同配件「{p.part_name}」", "contract_part", pid)
    await db.delete(p)
    await db.commit()
    return {"msg": "已删除"}


@router.get("/models")
async def get_contract_models(
    shop_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """获取合同配件的车型列表（去重），用于下拉选择"""
    query = select(ContractPart.vehicle_model).distinct()
    if shop_id:
        query = query.where(ContractPart.shop_id == shop_id)
    result = await db.execute(query)
    models = [r[0] for r in result if r[0]]
    return sorted(models)


@router.get("/template.xlsx")
async def download_template():
    """下载合同配件导入模板（xlsx格式）"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "服务器未安装 openpyxl，无法生成模板")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "合同配件模板"
    headers = ["车型", "配件名称", "配件型号", "配件价格", "工时费", "合计金额", "质保天数", "备注"]
    ws.append(headers)
    ws.append(["大众迈腾", "前刹车片", "A123", 280.0, 80.0, 360.0, 365, "原厂配件"])
    ws.append(["大众迈腾", "机油滤芯", "B456", 45.0, 30.0, 75.0, 180, "副厂"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import Response
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=contract-parts-template.xlsx"},
    )


@router.post("/import/xlsx")
async def import_contract_parts_xlsx(
    file: UploadFile = File(...),
    contract_id: int | None = None,
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_role("admin")),
):
    """从 xlsx 导入合同配件"""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "服务器未安装 openpyxl")

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content))
    ws = wb.active
    if ws is None:
        raise HTTPException(400, "Excel 文件为空")

    imported = 0
    skipped = 0
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[0]:
            skipped += 1
            continue
        # 列: 车型, 配件名称, 配件型号, 配件价格, 工时费, 合计金额, 质保天数, 备注
        vehicle_model = str(row[0] or "").strip()
        part_name = str(row[1] or "").strip()
        if not part_name:
            skipped += 1
            continue
        part_model = str(row[2] or "").strip()
        try:
            contract_price = float(row[3]) if row[3] is not None else 0.0
        except (ValueError, TypeError):
            contract_price = 0.0
        try:
            labor_fee = float(row[4]) if row[4] is not None else 0.0
        except (ValueError, TypeError):
            labor_fee = 0.0
        try:
            total_amount = float(row[5]) if row[5] is not None else 0.0
        except (ValueError, TypeError):
            total_amount = 0.0
        try:
            warranty_days = int(row[6]) if row[6] is not None else 0
        except (ValueError, TypeError):
            warranty_days = 0
        notes = str(row[7] or "").strip() if len(row) > 7 else ""

        p = ContractPart(
            contract_id=contract_id,
            vehicle_model=vehicle_model,
            part_name=part_name,
            part_model=part_model,
            contract_price=contract_price,
            labor_fee=labor_fee,
            total_amount=total_amount,
            warranty_days=warranty_days,
            notes=notes,
        )
        db.add(p)
        imported += 1
    await db.commit()
    await log_operation(db, user, f"批量导入了 {imported} 个合同配件（跳过 {skipped} 个）", "contract_part", 0)
    return {"imported": imported, "skipped": skipped}
